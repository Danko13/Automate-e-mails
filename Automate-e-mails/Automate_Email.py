# SENDS EMAILS BASED ON PAYMENT STATUS IN SPREADSHEET

import openpyxl, smtplib

# TO DO: Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
lastRow = sheet.max_row

latestMonth = sheet.cell(row=1, column=lastCol).value

# TO DO: Check each member's payment status.

unpaidMembers = {}
for r in range (2, lastRow + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# TO DO: Log in to email account.

conn = smtplib.SMTP('smtp.gmail.com', 587)
conn.ehlo()
conn.starttls()
conn.login('daniel.petrusevski13@gmail.com', 'daniel1303990')

# TO DO: Send out reminder emails
print (unpaidMembers)

for name, email in unpaidMembers.items():
    body = 'Subject: %s dues unpaid.\n\nDear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!' % (latestMonth, name, latestMonth)
    print ('Sending email to %s...' % email)
    sendmailStatus = conn.sendmail('daniel.petrusevski13@gmail.com', email, body)
    if sendmailStatus != {}:
        print ('There was a problem sending email to %s: %s' % (email, sendmailStatus))
conn.quit()